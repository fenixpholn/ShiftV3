import pandas as pd
import random
import json
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import PatternFill

class DutyScheduler:
    def __init__(self):
        self.doctors = []  # รายชื่อแพทย์
        self.roles = ["อช", "อญ", "สห", "ICU1", "ICU2"]  # หน้าที่เริ่มต้น
        self.holidays = []  # วันหยุดราชการ
        self.constraints = {}  # ข้อจำกัดของแพทย์แต่ละคน
        self.schedule = pd.DataFrame()  # ตารางเวร

    def add_doctor(self, name, max_shifts, unavailable_days, preferred_days, unavailable_roles):
        if name not in self.doctors:
            self.doctors.append(name)
        self.constraints[name] = {
            "max_shifts": max_shifts,
            "unavailable_days": unavailable_days,
            "preferred_days": preferred_days,
            "unavailable_roles": unavailable_roles
        }

    def set_holidays(self, holidays):
        self.holidays = holidays
    
    def generate_schedule(self, num_days):
        self.schedule = pd.DataFrame(index=range(1, num_days+1), columns=self.roles)
        duty_counts = {doc: 0 for doc in self.doctors}
        
        for day in range(1, num_days+1):
            available_doctors = [doc for doc in self.doctors if day not in self.constraints[doc]["unavailable_days"]]
            
            for role in self.roles:
                valid_doctors = [doc for doc in available_doctors if role not in self.constraints[doc]["unavailable_roles"] and duty_counts[doc] < self.constraints[doc]["max_shifts"]]
                
                if valid_doctors:
                    selected_doctor = random.choice(valid_doctors)
                    self.schedule.at[day, role] = selected_doctor
                    duty_counts[selected_doctor] += 1
                else:
                    self.schedule.at[day, role] = "-"
        
        return self.schedule

    def export_to_excel(self, filename):
        wb = Workbook()
        ws = wb.active
        ws.title = "Duty Schedule"
        
        ws.append(["Day"] + self.roles)
        for day, row in self.schedule.iterrows():
            ws.append([day] + list(row))
            if day in self.holidays:
                for col in range(1, len(self.roles)+2):
                    ws.cell(row=ws.max_row, column=col).fill = PatternFill(start_color="FFFF00", fill_type="solid")
        
        wb.save(filename)

# เริ่มต้น Streamlit App
st.title("ระบบจัดเวรแพทย์อัตโนมัติ")

# ใช้ session state เพื่อเก็บข้อมูลแพทย์
if "scheduler" not in st.session_state:
    st.session_state.scheduler = DutyScheduler()

scheduler = st.session_state.scheduler

num_days = st.number_input("จำนวนวันในเดือนนี้", min_value=1, max_value=31, value=30)
holidays = st.multiselect("เลือกวันหยุดราชการ", list(range(1, num_days+1)))
scheduler.set_holidays(holidays)

st.subheader("เพิ่มแพทย์")
new_doc_name = st.text_input("ชื่อแพทย์")
max_shifts = st.number_input("จำนวนเวรสูงสุดในเดือนนี้", min_value=1, max_value=num_days, value=10)
unavailable_days = st.multiselect("เลือกวันที่ไม่สะดวกอยู่เวร", list(range(1, num_days+1)))
preferred_days = st.multiselect("เลือกวันที่สะดวกอยู่เวร", list(range(1, num_days+1)))
unavailable_roles = st.multiselect("เลือกเวรที่ไม่สะดวก", scheduler.roles)

if st.button("เพิ่มแพทย์"):
    scheduler.add_doctor(new_doc_name, max_shifts, unavailable_days, preferred_days, unavailable_roles)
    st.success(f"เพิ่มแพทย์ {new_doc_name} เรียบร้อย")

st.subheader("รายชื่อแพทย์และเงื่อนไข")
for doctor in scheduler.doctors:
    if doctor in scheduler.constraints:
        with st.expander(f"{doctor} (สูงสุด {scheduler.constraints[doctor]['max_shifts']} เวร)"):
            new_max_shifts = st.number_input(f"ปรับจำนวนเวรสูงสุดของ {doctor}", min_value=1, max_value=num_days, value=scheduler.constraints[doctor]['max_shifts'])
            new_unavailable_days = st.multiselect(f"ปรับวันที่ {doctor} ไม่สะดวก", list(range(1, num_days+1)), default=scheduler.constraints[doctor]['unavailable_days'])
            new_preferred_days = st.multiselect(f"ปรับวันที่ {doctor} สะดวก", list(range(1, num_days+1)), default=scheduler.constraints[doctor]['preferred_days'])
            new_unavailable_roles = st.multiselect(f"ปรับเวรที่ {doctor} ไม่สะดวก", scheduler.roles, default=scheduler.constraints[doctor]['unavailable_roles'])
            
            if st.button(f"บันทึกการแก้ไขของ {doctor}"):
                scheduler.add_doctor(doctor, new_max_shifts, new_unavailable_days, new_preferred_days, new_unavailable_roles)
                st.success(f"บันทึกการแก้ไขของ {doctor} เรียบร้อย")

st.subheader("จัดตารางเวร")
if st.button("สร้างตารางเวร"):
    schedule = scheduler.generate_schedule(num_days)
    st.write(schedule)
    
    filename = "duty_schedule.xlsx"
    scheduler.export_to_excel(filename)
    with open(filename, "rb") as file:
        st.download_button(label="ดาวน์โหลดไฟล์ Excel", data=file, file_name=filename, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
